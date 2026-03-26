import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import messagebox
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Color
import xlwings as xw
import subprocess
import os
import sys
import msvcrt
import re
import tempfile
import shutil

# Absolute path to the directory where main.py is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# The root of your project (one level up from src/)
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

latex_escape_dict = {
    '#': '\\#',
    '$': '\\$',
    '%': '\\%',
    '&': '\\&',
    '_': '\\_',
    '{': '\\{',
    '}': '\\}',
    '~': '\\textasciitilde{}',
    '^': '\\textasciicircum{}',
    '\\': '\\textbackslash{}',
    '≥': '$\\ge$',
    '≤': '$\\le$',
    '≠': '$\\neq$',
    'α': '$\\alpha$',
    'β': '$\\beta$',
    'γ': '$\\gamma$',
    'δ': '$\\delta$',
    'ε': '$\\epsilon$',
    'ζ': '$\\zeta$',
    'η': '$\\eta$',
    'θ': '$\\theta$',
    'ι': '$\\iota$',
    'κ': '$\\kappa$',
    'λ': '$\\lambda$',
    'μ': '$\\mu$',
    'ν': '$\\nu$',
    'ξ': '$\\xi$',
    'ο': '$\\omicron$',
    'π': '$\\pi$',
    'ρ': '$\\rho$',
    'σ': '$\\sigma$',
    'τ': '$\\tau$',
    'υ': '$\\upsilon$',
    'φ': '$\\phi$',
    'χ': '$\\chi$',
    'ψ': '$\\psi$',
    'ω': '$\\omega$',
    'Α': 'A',
    'Β': 'B',
    'Γ': '$\\Gamma$',
    'Δ': '$\\Delta$',
    'Ε': 'E',
    'Ζ': 'Z',
    'Η': 'H',
    'Θ': '$\\Theta$',
    'Ι': 'I',
    'Κ': 'K',
    'Λ': '$\\Lambda$',
    'Μ': 'M',
    'Ν': 'N',
    'Ξ': '$\\Xi$',
    'Ο': 'O',
    'Π': '$\\Pi$',
    'Ρ': 'P',
    'Σ': '$\\Sigma$',
    'Τ': 'T',
    'Υ': '$\\Upsilon$',
    'Φ': '$\\Phi$',
    'Χ': 'X',
    'Ψ': '$\\Psi$',
    'Ω': '$\\Omega$',
    '\n':'\\\\',
    'V̇': '$\\dot{V}$',
    "ṁ": "$\\dot{m}$",
    "Q̇": "$\\dot{Q}$",
    "ẍ": "$\\ddot{x}$",
    '∙': '$\\cdot$'
}

def write_footer(default_text):
    """
    Shows a small window with a prompt and a single-line entry.
    Returns the text entered by the user.
    Exits program if window is closed.
    """
    input_window = tk.Toplevel()  # modal window
    input_window.title("Input")
    input_window.geometry("400x150")
    input_window.resizable(False, False)
    input_window.configure(bg="#f0f0f0")

    # Close handler: terminate program if user closes window
    def on_close():
        print("Window closed. Exiting...")
        input_window.destroy()
        sys.exit(0)

    input_window.protocol("WM_DELETE_WINDOW", on_close)

    # Variable to store result
    user_text = tk.StringVar(value=default_text)

    # Frame for layout
    frame = tk.Frame(input_window, bg="#f0f0f0")
    frame.pack(expand=True, padx=20, pady=20)

    # Prompt label
    label = tk.Label(frame, text="Write the footer:", font=("Arial", 12), bg="#f0f0f0")
    label.pack(pady=(0, 10))

    # Entry widget
    entry = tk.Entry(frame, textvariable=user_text, font=("Arial", 12), width=30)
    entry.pack(pady=(0, 15))
    entry.focus()

    # OK button
    def on_ok():
        input_window.quit()  # close mainloop

    style = ttk.Style()
    style.configure("TButton", font=("Arial", 11), padding=6)
    ok_button = ttk.Button(frame, text="OK", command=on_ok)
    ok_button.pack()

    # Make modal: wait for user input
    input_window.grab_set()
    input_window.mainloop()
    input_window.destroy()

    return user_text.get()

def select_gamma_or_machine():
    options = ["Gamma", "Single machine"]

    mode_window = tk.Tk()
    mode_window.title("Do you want a manual for a single machine or the whole gamma")

    mode_var = tk.StringVar(mode_window)
    mode_var.set(options[1])  # Default selection: "Gamma"

    def on_mode_select():
        mode_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        mode_window.destroy()
        sys.exit(0)

    mode_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(mode_window, text="Select mode:")
    label.pack(pady=10)

    menu = tk.OptionMenu(mode_window, mode_var, *options)
    menu.pack(pady=10)

    button = tk.Button(mode_window, text="OK", command=on_mode_select)
    button.pack(pady=10)

    mode_window.mainloop()
    mode_window.destroy()

    return mode_var.get()


def select_view():
    options = ["Manuale", "Preview"]

    mode_window = tk.Tk()
    mode_window.title("Select Mode")

    mode_var = tk.StringVar(mode_window)
    mode_var.set(options[0])  # Default selection: "Manuale"

    def on_mode_select():
        mode_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        mode_window.destroy()
        sys.exit(0)

    mode_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(mode_window, text="Select mode:")
    label.pack(pady=10)

    menu = tk.OptionMenu(mode_window, mode_var, *options)
    menu.pack(pady=10)

    button = tk.Button(mode_window, text="OK", command=on_mode_select)
    button.pack(pady=10)

    mode_window.mainloop()
    mode_window.destroy()

    return mode_var.get()


def select_sheet(sheet_names):
    sheet_window = tk.Tk()
    sheet_window.title("Select Sheet")

    sheet_menu = tk.StringVar(sheet_window)
    sheet_menu.set(sheet_names[1])  # default

    def on_sheet_select():
        selected_sheet = sheet_menu.get()
        sheet_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        sheet_window.destroy()
        sys.exit(0)

    sheet_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(sheet_window, text="Select a sheet:")
    label.pack(pady=10)
    menu = tk.OptionMenu(sheet_window, sheet_menu, *sheet_names)
    menu.pack(pady=10)
    button = tk.Button(sheet_window, text="OK", command=on_sheet_select)
    button.pack(pady=10)

    sheet_window.mainloop()
    sheet_window.destroy()

    return sheet_menu.get()


def get_sheet_names(wb):
    # Return the list of sheet names
    return wb.sheetnames


def read_excel(file_path):
    # make file into a workbook and get an active sheet
    wb = openpyxl.load_workbook(file_path)
    return wb


def collect_languages(sheet):
    languages = []
    for row in sheet.iter_rows(min_row=5):
        cell = row[6]
        if cell.value is None:
            return languages
        languages.append(cell.value)
    return languages


def collect_machines(sheet):
    start=False
    machines = dict()
    machines_rev = dict()
    for row in sheet.iter_rows(min_row=8):
        cell = row[5]
        if cell.value is None and start:
            break
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            if fill.fgColor.rgb == "FFD6DCE4":
                start=True
                name = cell.value
                number = row[6].value
                machines[name] = number
                machines_rev[number] = name
    return machines, machines_rev


def create_groups(sheet):
    groups = dict()
    for merged_range in sheet.merged_cells.ranges:
        if 'M7' in merged_range:
            mincol = merged_range.min_col - 1
            maxcol = merged_range.max_col
    for i in range(mincol, maxcol + 1):
        for j in range(9, sheet.max_row + 1):
            cell = sheet.cell(row=j, column=i)
            value = cell.value
            if value is None or str(value).strip() == "":
                break
            if j == 9:
                key = value
                groups[key] = []
            elif j > 9:
                groups[key].append(value)
    return groups

def create_gamma_codes(sheet):
    codes=dict()
    data=list(sheet.iter_rows(min_row=2,max_row=5,min_col=13,max_col=16,values_only=True))
    headers=[str(h).lower() for h in data[0][1:]]
    for row in data[1:]:
        row_name=str(row[0]).lower()
        for col_name, value in zip(headers,row[1:]):
            codes[(row_name,col_name)]=value
    return codes


def select_language(languages):
    # select the language which you want to see out of a menu

    language_window = tk.Tk()
    language_window.title("Select Language")

    language_menu = tk.StringVar(language_window)
    language_menu.set(languages[0])  # default is italian

    def on_language_select():
        language_window.quit()  # Close the menu

    def on_close():
        print("Window closed. Exiting...")
        language_window.destroy()
        sys.exit(0)

    language_window.protocol("WM_DELETE_WINDOW", on_close)
    # create dropdown menu

    label = tk.Label(language_window, text="Select a language:")
    label.pack(pady=10)
    menu = tk.OptionMenu(language_window, language_menu, *languages)
    menu.pack(pady=10)
    button = tk.Button(language_window, text="Select Language", command=on_language_select)
    button.pack(pady=10)
    language_window.mainloop()
    language_window.destroy()
    return language_menu.get()


def create_gamma_dict(groups):
    gammas = dict()

    gammas["P"] = groups["P"]
    gammas["R"] = groups["R"]
    gammas["D"] = groups["D"]
    return gammas


def select_gamma(gammaDict):
    # opens a menu to select machine, showing only rows relevant to that machine

    gammas = [values[0] for values in gammaDict.values() if values]
    selected_gamma = {"value": ""}

    def on_gamma_select():
        selected_gamma["value"] = gamma_menu.get()
        gamma_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        gamma_window.destroy()
        sys.exit(0)

    gamma_window = tk.Tk()
    gamma_window.title("Select Gamma")
    gamma_window.protocol("WM_DELETE_WINDOW", on_close)

    gamma_menu = tk.StringVar(gamma_window)
    gamma_menu.set(gammas[0])  # default

    label = tk.Label(gamma_window, text="Select a gamma:")
    label.pack(pady=10)
    menu = tk.OptionMenu(gamma_window, gamma_menu, *gammas)
    menu.pack(pady=10)
    button = tk.Button(gamma_window, text="OK", command=on_gamma_select)
    button.pack(pady=10)

    gamma_window.mainloop()
    gamma_window.destroy()
    for k in gammaDict.keys():
        if selected_gamma["value"] in gammaDict[k]:
            return k
    return selected_gamma["value"]


def create_version_dict(groups):
    versions = dict()
    versions["W"] = groups["W"]
    versions["A"] = groups["A"]
    versions["G"] = groups["G"]
    return versions


def select_version(versionDict):
    # opens a menu to select machine, showing only rows relevant to that machine

    versions = [values[0] for values in versionDict.values() if values]

    version_window = tk.Tk()
    version_window.title("Select Version")

    version_menu = tk.StringVar(version_window)
    version_menu.set(versions[0])  # default

    selected_version = {"value": ""}

    def on_version_select():
        selected_version["value"] = version_menu.get()
        version_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        version_window.destroy()
        sys.exit(0)

    version_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(version_window, text="Select a version:")
    label.pack(pady=10)
    menu = tk.OptionMenu(version_window, version_menu, *versions)
    menu.pack(pady=10)
    button = tk.Button(version_window, text="OK", command=on_version_select)
    button.pack(pady=10)

    version_window.mainloop()
    version_window.destroy()

    for k in versionDict.keys():
        if selected_version["value"] in versionDict[k]:
            return k
    return versionDict.keys()[0]


def select_machine(availableMachines, machines, machines_rev):
    # opens a menu to select machine, showing only rows relevant to that machine
    machineList = list(machines_rev[m] for m in availableMachines)

    machine_window = tk.Tk()
    machine_window.title("Select Machine")

    machine_menu = tk.StringVar(machine_window)
    machine_menu.set(machineList[0])  # default

    def on_machine_select():
        selected_machine = machine_menu.get()
        machine_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        machine_window.destroy()
        sys.exit(0)

    machine_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(machine_window, text="Select a machine:")
    label.pack(pady=10)
    menu = tk.OptionMenu(machine_window, machine_menu, *machineList)
    menu.pack(pady=10)
    button = tk.Button(machine_window, text="OK", command=on_machine_select)
    button.pack(pady=10)

    machine_window.mainloop()
    machine_window.destroy()

    return machines[machine_menu.get()]


def create_selected_groups(machine, groups):
    selected_groups = []
    for g in groups.keys():
        if machine in groups[g]:
            selected_groups.append(g)
    return selected_groups


def select_mode():
    machine_window = tk.Tk()
    machine_window.title("Select mode")

    # Options and their corresponding return values
    options = {
        "Reversible": "RT",
        "Non-reversible": "T"
    }

    machine_menu = tk.StringVar(machine_window)
    machine_menu.set("Reversible")  # Default selection

    def on_mode_select():
        machine_window.quit()

    def on_close():
        print("Window closed. Exiting...")
        machine_window.destroy()
        sys.exit(0)

    machine_window.protocol("WM_DELETE_WINDOW", on_close)

    label = tk.Label(machine_window, text="Select the mode:")
    label.pack(pady=10)

    menu = tk.OptionMenu(machine_window, machine_menu, *options.keys())
    menu.pack(pady=10)

    button = tk.Button(machine_window, text="OK", command=on_mode_select)
    button.pack(pady=10)

    machine_window.mainloop()
    machine_window.destroy()

    return options[machine_menu.get()]


def format_title(sheet,interface_sheet,sheet_xlw, column_letter):
    cell_address = f"{column_letter}3"
    cell=sheet[cell_address]
    size=int(cell.font.sz)
    hor= interface_sheet["B7"].value
    ver= interface_sheet["C7"].value

    title=f"\\vspace*{{{ver}cm}}\\hspace*{{{hor}cm}}{{\\fontsize{{{size}}}{{{size+4}}}"
    title+=latex_formatting(sheet_xlw,cell_address)
    title+= "}"


    return title

def format_subtitle(machine,interface_sheet, mode):
    if mode=="T":
        machine = machine.replace("(R)", "", 1)
    else:
        machine= machine.replace("(", "").replace(")", "")

    hor=interface_sheet["C10"].value
    ver=interface_sheet["C11"].value
    cell=interface_sheet["C12"]
    size=cell.font.sz
    color=""
    if cell.font.color and cell.font.color.type == "rgb":
        hex_color = cell.font.color.rgb  # e.g., 'FF112233'

        # Remove alpha if present (first two chars)
        hex_color = hex_color[-6:]  # take last 6 chars

        # Convert to tuple of integers
        rgb = tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
        if isinstance(rgb, tuple) and len(rgb) == 3:
            color = f"\\textcolor[RGB]{{{rgb[0]},{rgb[1]},{rgb[2]}}}"

    subtitle=f"""\\vspace*{{{ver}cm}}\\hspace*{{{hor}cm}}{{\\fontsize{{{size}}}{{{size+3}}}{color}{{\\textbf{{{machine}}}}}}}"""
    return subtitle



def write_packages(sheet, interface_sheet,sheet_xlw, column, machine,mode, footer):
    column_letter = get_column_letter(column + 1)

    content = r"""\documentclass[11pt,twoside]{report}
    \usepackage{carlito}
    \usepackage{titlesec}      % for custom headings
    \usepackage{lipsum}
    \titleformat{\chapter}[block]       
    {\bfseries\fontsize{22}{26}\selectfont} 
    {}                      
    {10pt}                               
    {}                                   
    [\vspace{2pt}\titlerule] 
    \usepackage{graphicx} % Required for inserting images
    \usepackage[italian]{babel}
    \usepackage[utf8]{inputenc}
    \usepackage{amsmath}
    \usepackage{amssymb}
    \usepackage{amsthm}
    \usepackage{algorithm}
    \usepackage{xcolor}
    \usepackage[table]{xcolor}
    \usepackage{tikz}
    \usepackage[T1]{fontenc}
    \usepackage{pifont}
    \usepackage{caption}
    \usepackage{subcaption}
    \usepackage{fancyhdr}
    \usepackage{titlesec}
    \usepackage{blindtext}
    \usepackage{hyperref}
    \usepackage{float}
    \usepackage[a4paper, hmargin=2cm, vmargin=2cm]{geometry}
    \usepackage{array}
    \usepackage{colortbl}
    \usepackage{multirow}
    \usepackage{eso-pic}
    \usepackage{anyfontsize}
    \usepackage{changepage}
    \usepackage{makecell}
    \raggedbottom
   
\pagestyle{fancy}
\fancyhf{} % clear all header and footer fields

% footer: centered text
\fancyfoot[C]{"""
    content+=footer

    content+=r"""}
\fancyfoot[RO,LE]{\thepage}  

% remove the default header line
\renewcommand{\headrulewidth}{0pt}
% keep or adjust the footer line thickness (0pt = none)
\renewcommand{\footrulewidth}{0.4pt}

\fancypagestyle{plain}{%
  \fancyhf{} % clear default
  \fancyfoot[C]{"""

    content+=footer

    content+=r"""}
  \fancyfoot[RO,LE]{\thepage}
  \renewcommand{\headrulewidth}{0pt}
  \renewcommand{\footrulewidth}{0.4pt}
}
   
    \date{}
    
    % Command to add background image
\newcommand\CoverBackground{%
  \AddToShipoutPictureBG*{%
    \AtPageLowerLeft{%
      \includegraphics[width=\paperwidth,height=\paperheight]{"""

    cell = sheet[f"{column_letter}1"]
    image_path = resource_path(os.path.join("images", str(cell.value))).replace("\\", "/")
    content += image_path
    content+=r"""}%
    }%
  }%
} 
    \setlength{\fboxsep}{10pt}
    \renewcommand{\arraystretch}{1.3}
    \newcolumntype{C}[1]{>{\centering\arraybackslash}m{#1}}
    \setcounter{secnumdepth}{0}
    
    

    \titlespacing*{\chapter}{0pt}{0pt}{0pt}
    \titlespacing*{\section}{0pt}{5pt}{5pt}
    \setlength{\parskip}{10pt}   
    \setlength{\parindent}{0pt}

    \begin{document}

    
\CoverBackground
    \begin{titlepage}""" +"\n"
    content+=format_title(sheet,interface_sheet, sheet_xlw, column_letter) + "\n"
    content+="\n"
    content+=format_subtitle(machine,interface_sheet, mode) + "\n"
    content+=r"""\end{titlepage}
    
    \ClearShipoutPictureBG

    \tableofcontents"""
    return content


def is_match(groups, condition, machine, machines_rev):
    # Handle universal condition (always matches)
    if condition:
        if condition == "ALL":
            return True

        # Split condition by commas and check each part
        parts = [part.strip() for part in condition.split(',')]

        for part in parts:
            if part in groups:
                return True
            elif machines_rev[machine] == part:
                return True
    return False


def collect_cell_contents(column, selected_groups, sheet, sheet_xlw, selected_machine, machines_rev, wb, wb_xlw, view):
    content = ""
    previousbg = ""
    color = False
    previous_chapter_row = None
    something_between = False
    total_rows = sheet.max_row - 1

    progress_win = tk.Toplevel()
    progress_win.title("Progress")

    progress = ttk.Progressbar(progress_win, orient="horizontal", length=300, mode="determinate")
    progress.grid(row=0, column=0, padx=10, pady=10)
    progress["maximum"] = total_rows

    label = tk.Label(progress_win, text="Processing...")
    label.grid(row=1, column=0, padx=10, pady=5)

    progress_win.update_idletasks()

    for i, row in enumerate(sheet.iter_rows(min_row=3)):

        progress["value"] = i
        label.config(text=f"Processing row {i + 1} of {total_rows}")
        progress_win.update()
        progress_win.update_idletasks()

        column_letter = get_column_letter(column + 1)
        cell_address = f"{column_letter}{row[0].row}"
        if is_match(selected_groups, row[0].value, selected_machine, machines_rev):
            cell = row[column]
            font_color = cell.font.color
            cell_value = cell.value
            fill = cell.fill
            if cell_value is None or cell_value=="":
                continue
            if fill and fill.fgColor and fill.fgColor.type == "rgb" and fill.fgColor.rgb != "00000000":
                bg_rgb = fill.fgColor.rgb[2:] if fill.fgColor.rgb.startswith("FF") else fill.fgColor.rgb
                bg = f"[HTML]{{{bg_rgb}}}"

                if bg != previousbg:
                    if color == True:
                        content += "\\end{minipage}}\\end{center}"
                    content += "\\begin{center}"
                    content += f"\\fcolorbox{{black}}{bg}{{\\begin{{minipage}}{{0.9\\textwidth}}{{"

                color = True
                alignment = cell.alignment.horizontal

                if view == "Preview":
                    if alignment=="center":
                        content+="\\begin{center}"
                    content += f" {{\\color{{red}} \\textbf{{\\Large {cell_address}}}}} {{{latex_formatting(sheet_xlw, cell_address)}}} \\\\"
                    if alignment=="center":
                        content+="\\end{center}"
                else:
                    if alignment=="center":
                        content+="\\begin{center}"
                    content += f" {{{latex_formatting(sheet_xlw, cell_address)}}} \\\\"
                    if alignment=="center":
                        content+="\\end{center}"

                previousbg = bg
                something_between = True


            else:
                if color == True:
                    content += "}\\end{minipage}}\\end{center}"
                    color = False
                    previousbg = None

                if isinstance(font_color, Color):
                    font_color_value = font_color.rgb
                    if font_color_value not in {
    "FF189EDA",  # chapter
    "FFFFC000",  # table
    "FF595959",  # section
    "FF15A8DD",  # subsection
    "FF2F75B5",  # subsubsection
    "FF7030A0",  # image
                    }:
                        something_between = True
                        alignment = cell.alignment.horizontal
                        if view == "Preview":
                            content += f"{{\\color{{red}} \\textbf{{\\Large {cell_address}}}}}\n"
                        if alignment == "center":
                            content += "\\begin{center}" + latex_formatting(sheet_xlw,
                                                                            cell_address) + r"\par" + "\\end{center}" + "\n"
                        else:
                            content += latex_formatting(sheet_xlw, cell_address) + r"\par" + "\n"

                    elif font_color_value == "FF189EDA":
                        if previous_chapter_row is not None and not something_between:
                            content += "\\textbf{N/A} \\\\ \n"

                        content += "\\chapter*{"
                        if view == "Preview":
                            content += cell_address
                        chapter= f" {str(latex_formatting(sheet_xlw, cell_address))}}}" + "\n"
                        content+=chapter
                        content+= f"\\addcontentsline{{toc}}{{chapter}}{{{cell_value}}}"

                        previous_chapter_row = row[0].row
                        something_between = False

                    elif font_color_value == "FFFFC000":
                        something_between = True
                        parts = [part.strip("=$") for part in cell_value.split("!")]
                        table_sheet = wb[parts[0]]
                        table_sheet_xlw = wb_xlw.sheets[parts[0]]
                        cell_range = parts[1]
                        if view == "Preview":
                            content += f"{{\\color{{red}} \\textbf{{\\Large {cell_address}}}}}\n"
                        content += format_table(table_sheet, cell_range, column_letter, table_sheet_xlw) + "\n"
                    elif font_color_value == "FF595959":
                        something_between = True
                        heading = ""
                        if view == "Preview":
                            heading += cell_address + " "
                        heading += latex_formatting(sheet_xlw, cell_address)
                        content += f"\\section{{{str(heading)}}}" + "\n"
                    elif font_color_value == "FF15A8DD":
                        something_between = True
                        subheading = ""
                        if view == "Preview":
                            subheading += cell_address + " "
                        subheading += latex_formatting(sheet_xlw, cell_address)
                        content += f"\\subsection{{{str(subheading)}}}" + "\n"
                    elif font_color_value == "FF2F75B5":
                        something_between = True
                        subsubheading = ""
                        if view == "Preview":
                            subsubheading += cell_address + " "
                        subsubheading += latex_formatting(sheet_xlw, cell_address)
                        content += f"\\subsubsection{{{str(subsubheading)}}}" + "\n"
                    elif font_color_value == "FF7030A0":
                        something_between = True
                        if view == "Preview":
                            content += f"{{\\color{{red}} \\textbf{{\\Large {cell_address}}}}}\n"
                        image = "\\begin{figure}[H]\n"
                        image += "\\centering\n"
                        image_raw_path = os.path.join(PROJECT_ROOT, "images", str(cell_value))
                        image_path = image_raw_path.replace("\\", "/")
                        image += f"\\includegraphics[width=0.95\\textwidth]{{{image_path}}}\n"
                        image += "\\end{figure}\n"
                        content += image
                    else:
                        something_between = True
                        alignment=cell.alignment.horizontal
                        if view == "Preview":
                            content += f"{{\\color{{red}} \\textbf{{\\Large {cell_address}}}}}\n"
                        if alignment == "center":
                            content += "\\begin{center}" + latex_formatting(sheet_xlw, cell_address) + r"\par" + "\\end{center}" + "\n"
                        else:
                            content += latex_formatting(sheet_xlw, cell_address) + r"\par" + "\n"
                else:
                    something_between = True
                    alignment = cell.alignment.horizontal
                    if view == "Preview":
                        content += f"{{\\color{{red}} \\textbf{{\\Large {cell_address}}}}}\n"
                    if alignment == "center":
                        content += "\\begin{center}" + latex_formatting(sheet_xlw,
                                                                        cell_address) + r"\par" + "\\end{center}" + "\n"
                    else:
                        content += latex_formatting(sheet_xlw, cell_address) + r"\par" + "\n"

    label.config(text="Done!")
    progress_win.update()
    progress_win.after(1000, progress_win.destroy)

    return content


def format_table(sheet, cell_range, language, sheet_xlw):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    col_count = max_col - min_col + 1

    merged_ranges = sheet.merged_cells.ranges
    filtered_merged_ranges = []
    for m_range in merged_ranges:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(m_range))
        if (
                m_max_row >= min_row and m_min_row <= max_row and
                m_max_col >= min_col and m_min_col <= max_col
        ):
            filtered_merged_ranges.append(m_range)

    merged_map = {}
    for m_range in filtered_merged_ranges:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(m_range))
        for r in range(m_min_row, m_max_row + 1):
            for c in range(m_min_col, m_max_col + 1):
                merged_map[(r, c)] = (m_min_row, m_min_col, m_max_row, m_max_col)

        # Get actual column widths
    col_widths_cm = []
    for col in range(min_col, max_col + 1):
        col_letter = get_column_letter(col)
        width = sheet.column_dimensions[col_letter].width
        col_widths_cm.append(excel_colwidth_to_cm(width))
    sum_width = sum(col_widths_cm, 0)
    #if sum_width > 15.5:
    #    for i in range(0, len(col_widths_cm)):
    #        col_widths_cm[i] = col_widths_cm[i] / sum_width * 0.6
    #    latex = "\\begin{center}"
    #    latex += "\\begin{tabular}{|" + " | ".join(
    #        [f"C{{{w:.2f}\\textwidth}}" for w in col_widths_cm]) + "|}\n\\hline\n"
    #else:
    latex=r"""\begin{table}[H]
            \begin{adjustwidth}{-1.5cm}{-1.5cm}
            \begin{center}"""
    latex += "\\begin{tabular}{|" + " | ".join([f"C{{{w:.2f}cm}}" for w in col_widths_cm]) + "|}\n\\hline\n"

    for row_idx in range(min_row, max_row + 1):
        row_cells = []

        for col_idx in range(min_col, max_col + 1):
            coord = (row_idx, col_idx)

            if coord in merged_map:
                m_min_row, m_min_col, m_max_row, m_max_col = merged_map[coord]
                rowspan = m_max_row - m_min_row + 1
                colspan = m_max_col - m_min_col + 1

                if m_min_row != m_max_row and m_min_col == m_max_col:
                    # Pure vertical span
                    if row_idx < m_max_row:
                        # In upper part of merge: insert colored empty cell
                        origin_cell = sheet.cell(row=m_min_row, column=m_min_col)
                        bg = ""
                        fill = origin_cell.fill
                        if fill and fill.fgColor and fill.fgColor.type == "rgb":
                            bg_rgb = fill.fgColor.rgb[2:] if fill.fgColor.rgb.startswith("FF") else fill.fgColor.rgb
                            if bg_rgb != "00000000":
                                bg = f"\\cellcolor[HTML]{{{bg_rgb}}}"
                        row_cells.append(bg)
                        continue
                    elif row_idx == m_max_row:
                        # Bottom of merge: insert multirow{-n} with content
                        cell = sheet.cell(row=m_min_row, column=m_min_col)
                        text = latex_formatting(sheet_xlw, cell.coordinate,table=True)
                        if cell.font.sz:
                            text = f"\\fontsize{{{int(cell.font.sz)-2}}}{{{int(cell.font.sz)}}}\\selectfont \\begin{{center}} {text} \\end{{center}}"

                        bg = ""
                        fill = cell.fill
                        if fill and fill.fgColor and fill.fgColor.type == "rgb":
                            bg_rgb = fill.fgColor.rgb[2:] if fill.fgColor.rgb.startswith("FF") else fill.fgColor.rgb
                            if bg_rgb != "00000000":
                                bg = f"\\cellcolor[HTML]{{{bg_rgb}}}"
                        col_idx = m_min_col - min_col  # index in col_widths_cm
                        col_width = col_widths_cm[col_idx]
                        cell_latex = f"\\multirow{{-{rowspan}}}{{{col_width:.2f}cm}}{{\\strut {bg}{text}}}"
                        row_cells.append(cell_latex)
                        continue
                    else:
                        continue
                elif coord != (m_min_row, m_min_col):
                    continue
                else:
                    # Horizontal or full merge — process normally
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    text = latex_formatting(sheet_xlw, cell.coordinate,table=True)
                    if cell.font.sz:
                        text = f"\\fontsize{{{int(cell.font.sz)-2}}}{{{int(cell.font.sz)}}}\\selectfont {text}"

                    bg = ""
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.type == "rgb":
                        bg_rgb = fill.fgColor.rgb[2:] if fill.fgColor.rgb.startswith("FF") else fill.fgColor.rgb
                        if bg_rgb != "00000000":
                            bg = f"\\cellcolor[HTML]{{{bg_rgb}}}"

                    cell_latex = f"\\strut {bg}{text}"

                    if colspan > 1:
                        total_width = sum(col_widths_cm[col_idx - min_col: col_idx - min_col + colspan])
                        #if sum_width > 15.5:
                        #    cell_latex = f"\\multicolumn{{{colspan}}}{{|C{{\\dimexpr {total_width:.2f}\\textwidth+{2 * (colspan - 1)}\\tabcolsep+{colspan - 2}\\arrayrulewidth\\relax}}|}}{{{cell_latex}}}"
                        #else:
                        cell_latex = f"\\multicolumn{{{colspan}}}{{|C{{\\dimexpr {total_width:.2f}cm+{2 * (colspan - 1)}\\tabcolsep+{colspan - 2}\\arrayrulewidth\\relax}}|}}{{{cell_latex}}}"
                    row_cells.append(cell_latex)
            else:
                # Normal (non-merged) cell
                cell = sheet.cell(row=row_idx, column=col_idx)
                text = latex_formatting(sheet_xlw, cell.coordinate,table=True)
                if cell.font.sz:
                    text = f"\\fontsize{{{int(cell.font.sz)-2}}}{{{int(cell.font.sz)}}}\\selectfont {text}"

                bg = ""
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.type == "rgb":
                    bg_rgb = fill.fgColor.rgb[2:] if fill.fgColor.rgb.startswith("FF") else fill.fgColor.rgb
                    if bg_rgb != "00000000":
                        bg = f"\\cellcolor[HTML]{{{bg_rgb}}}"

                cell_latex = f"\\strut {bg}{text}"
                row_cells.append(cell_latex)

        latex += " & ".join(row_cells) + " \\\\\n\\hline\n"

    latex += "\\end{tabular}"
    latex += "\\end{center}"
    latex += r"""\end{adjustwidth}
\end{table}"""
    return latex


def excel_colwidth_to_cm(width):
    if width is None:
        width = 8.43
    pixels = int((256 * width + int(128 / 7)) / 256 * 7)
    return pixels * 0.026458333 * 0.9

def latex_formatting(sheet_xlw, cell_address, table=False):
    def escape_latex_special_chars(s, dict):
        return "".join(dict.get(c,c) for c in s)

    cell = sheet_xlw.range(cell_address)
    text = cell.value

    if text is None or (isinstance(text, str) and text.strip() == ''):
        return ""
    if not isinstance(text, str):
        if hasattr(text, "is_integer") and text.is_integer():
            text = int(text)
        text = str(text)

    font = cell.font
    try:
        fontt = font.api._inner
        underline_num= fontt.Underline
        if underline_num==(2 or 4):
            underline=True
        else:
            underline = False
    except Exception:
        underline=False

    color_val = None
    try:
        c = font.color
        if isinstance(c, tuple) and len(c) == 3:
               color_val= f"[RGB]{{{c[0]},{c[1]},{c[2]}}}"

    except Exception:
        pass

    global_styles = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": underline,
        "superscript": fontt.Superscript,
        "subscript": fontt.Subscript,
        "color": color_val,
    }

    # Wrap helper for whole-cell styles
    def wrap_whole(text, styles):
        styled = text
        if table:
            if r"\\" in styled:
                styled=f"\\makecell{{{styled}}}"
        if styles["bold"]:
            styled = r"\textbf{" + styled + "}"
        if styles["italic"]:
            styled = r"\textit{" + styled + "}"
        if styles["underline"]:
            styled = r"\underline{" + styled + "}"
        if styles["color"]:
            styled = r"\textcolor{}{{{}}}".format(styles["color"], styled)
        return styled

    if all(v is not None for v in global_styles.values()):
        text = escape_latex_special_chars(text, latex_escape_dict)
        return wrap_whole(text, global_styles)

    result = ""
    current_style = {k: None for k in global_styles}
    buffer = ""

    def flush_buffer(style, buffer):
        if not buffer:
            return ""
        styled = buffer
        if style["bold"]:
            styled = r"\textbf{" + styled + "}"
        if style["italic"]:
            styled = r"\textit{" + styled + "}"
        if style["underline"]:
            styled = r"\underline{" + styled + "}"
        if style["color"]:
            styled = r"\textcolor" + style['color'] + "{" + styled + "}"
        if style["superscript"]:
            styled = r"$^{" + styled + "}$"
        if style["subscript"]:
            styled = r"$_{" + styled + "}$"
        return styled
    #try:
    chars=cell.characters
    for i, char in enumerate(text):
        char=escape_latex_special_chars(char,latex_escape_dict)

        ch_fmt=chars[i].font
        ch_fmtt=ch_fmt.api._inner

        bold = global_styles["bold"] if global_styles["bold"] is not None else bool(ch_fmt.bold)
        italic = global_styles["italic"] if global_styles["italic"] is not None else bool(ch_fmt.italic)
        underline_num_ch= ch_fmtt.Underline
        if underline_num_ch==(2 or 4):
            underline_ch=True
        else:
            underline_ch = False
        underline = global_styles["underline"] if global_styles["underline"] is not None else bool(underline_ch)
        superscript = global_styles["superscript"] if global_styles["superscript"] is not None else bool(ch_fmtt.Superscript)
        subscript = global_styles["subscript"] if global_styles["subscript"] is not None else bool(ch_fmtt.Subscript)

        if global_styles["color"] is not None:
            color = global_styles["color"]
        else:
            color_tuple = getattr(ch_fmt, "color", None)
            color = None
            if isinstance(color_tuple, tuple) and len(color_tuple) == 3 and color_tuple != (0, 0, 0):
                color = f"[RGB]{{{color_tuple[0]},{color_tuple[1]},{color_tuple[2]}}}"

        new_style = {
            "bold": bold,
            "italic": italic,
            "underline": underline,
            "color": color,
            "superscript": superscript,
            "subscript": subscript
        }

        if new_style != current_style:
            result += flush_buffer(current_style, buffer)
            buffer = ""
            current_style = new_style
        buffer += char

    result += flush_buffer(current_style, buffer)

    #except Exception:
    #    text=escape_latex_special_chars(text, latex_escape_dict)
    #    return wrap_whole(text,global_styles)


    return result


def latex_to_pdf_with_dialog(latex_code, default_name):
    # Tk root
    root = tk.Tk()
    root.withdraw()

    save_path = filedialog.asksaveasfilename(
        initialfile=f"{default_name}.pdf",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        title="Save PDF as"
    )
    if not save_path:
        print("Save cancelled.")
        return

    tempdir = os.getcwd()
    tex_file_path = os.path.join(tempdir, "document.tex")
    with open(tex_file_path, "w", encoding="utf-8") as f:
        f.write(latex_code)

    try:
        for i in range(2):
            result = subprocess.run(
                ["pdflatex", "-interaction=nonstopmode", "document.tex"],
                cwd=tempdir,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )

            log_file_path = os.path.join(tempdir, "document.log")
            if not os.path.exists(log_file_path):
                messagebox.showerror("Error", "Log file not found.")
                return
            else:
                errors_warnings_bb = []
                with open(log_file_path, "r", encoding="utf-8", errors="ignore") as log:
                    lines_to_grab = 0
                    for line in log:
                        # Capture errors
                        if line.startswith("!"):
                            errors_warnings_bb.append("-" * 20)  # Separator
                            errors_warnings_bb.append("Error: " + line.strip())
                            lines_to_grab = 10  # Set the counter to grab context
                            continue

                        # Grab context lines if the counter is active
                        if lines_to_grab > 0:
                            errors_warnings_bb.append("       " + line.strip())
                            lines_to_grab -= 1
                            continue
                        # Capture LaTeX warnings
                        elif "Warning" in line:
                            errors_warnings_bb.append("Warning: " + line.strip())


                if errors_warnings_bb:
                    msg_file = os.path.splitext(save_path)[0] + "_messages.txt"
                    with open(msg_file, "w", encoding="utf-8") as f:
                        f.write("\n".join(errors_warnings_bb))
                    print(f"Errors, warnings, saved to {msg_file}")

            if result.returncode != 0:
                messagebox.showerror("Compilation Error", "LaTeX compilation failed. Check console output.")
                return

    except Exception as e:
        messagebox.showerror("Unexpected Error", str(e))
        return

    # Move PDF
    generated_pdf = os.path.join(tempdir, "document.pdf")
    if os.path.exists(generated_pdf):
        os.replace(generated_pdf, save_path)
        messagebox.showinfo("Success", f"PDF saved to:\n{save_path}")
    else:
        messagebox.showerror("Error", "PDF was not generated.")


def write_and_recalculate(interface_sheet, table_sheet, language, machine, mode):
    interface_sheet["B3"].value = machine
    interface_sheet["C3"].value = mode
    interface_sheet["D3"].value = language
    table_sheet.api.Calculate()


def resource_path(relative_path):
    """ Get the absolute path to the resource, works for development and packaging. """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def main():
    lock_file_path = os.path.join(os.getenv("TEMP"), "my_program.lock")

    try:
        lock_file = open(lock_file_path, 'w')
        msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
    except OSError:
        print("Another instance is already running.")
        sys.exit(0)

    server_file_path = os.path.join(PROJECT_ROOT, "data", "Manuale esempio.xlsm")
    temp_dir = tempfile.gettempdir()
    local_copy_path = os.path.join(temp_dir, "Manuale_temp_copy.xlsm")
    shutil.copy2(server_file_path, local_copy_path)

    file_path = local_copy_path
    wb = read_excel(file_path)
    sheet_names = get_sheet_names(wb)
    interface_sheet = wb["Interfaccia di selezione"]
    languages = collect_languages(interface_sheet)

    machines, machines_rev = collect_machines(interface_sheet)
    groups = create_groups(interface_sheet)

    global root
    root = tk.Tk()
    root.withdraw()
    selected_sheet = select_sheet(sheet_names)
    sheet = wb[selected_sheet]
    language = select_language(languages)
    column = languages.index(language) + 1
    versions = create_version_dict(groups)
    selected_version = groups[select_version(versions)]
    vers=selected_version[0].lower()
    b=select_gamma_or_machine()
    gammas = create_gamma_dict(groups)
    selected_gammas = groups[select_gamma(gammas)]
    gam=selected_gammas[0].lower()
    mode = select_mode()
    if b=='Single machine':
        machines_for_menu = list(set(selected_version) & set(selected_gammas))
        if not machines_for_menu:
            root = tk.Tk()
            root.withdraw()
            # Show error dialog
            messagebox.showerror("Error", "No such machine, please try again.")
            sys.exit()
        selected_machine = select_machine(machines_for_menu, machines, machines_rev)
        selected_groups = create_selected_groups(selected_machine, groups)
        selected_groups.append(mode)
    else:
        gamma_codes=create_gamma_codes(interface_sheet)
        print(gamma_codes)
        selected_groups=gamma_codes[(gam,vers)]
        selected_machine=1
    view = select_view()
    footer_temp=write_footer(sheet[f"{get_column_letter(column+1)}2"].value)
    footer="".join(latex_escape_dict.get(c,c) for c in footer_temp)
    app = xw.App(visible=False)
    wb_xlw = app.books.open(file_path)
    try:
        sheet_xlw = wb_xlw.sheets[selected_sheet]
        interface_sheet_xlw = wb_xlw.sheets["Interfaccia di selezione"]
        table_sheet_xlw = wb_xlw.sheets["Tabelle"]
        write_and_recalculate(interface_sheet_xlw, table_sheet_xlw, language, machines_rev[selected_machine], mode)
        wb_xlw.save(file_path)
        content = write_packages(sheet, interface_sheet, sheet_xlw, column, machines_rev[selected_machine], mode, footer)
        content += "\n"
        content += collect_cell_contents(column, selected_groups, sheet, sheet_xlw, selected_machine, machines_rev, wb,
                                         wb_xlw, view)
        content += r"""\end{document}"""

        wb_xlw.save(file_path)
        latex_to_pdf_with_dialog(content, footer_temp)

    finally:
        wb_xlw.close()
        app.quit()
        try:
            os.remove(local_copy_path)
        except Exception as e:
            print(f"Warning could not delete temp file: {e}")

if __name__ == "__main__":
    main()